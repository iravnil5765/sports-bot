"""
Generates an Excel tracker that matches the SPR-style spreadsheet.
Columns: DATE | SPORT TYPE | BET TYPE | ODDS | UNITS BET | BET AMOUNT | RESULT | CHANGE IN $ | BANKROLL
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import database
import config

TRACKER_PATH = os.path.join(os.path.dirname(__file__), 'picks_tracker.xlsx')

# ── Colours ────────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)

HEADER_FILL   = _fill('CC00CC')   # purple like the screenshot
HEADER_FONT   = _font(bold=True, color='FFFFFF', size=10)

SPORT_COLORS = {
    'NFL':      ('FF4500', 'FFFFFF'),
    'NBA':      ('FF8C00', 'FFFFFF'),
    'NHL':      ('1E90FF', 'FFFFFF'),
    'MLB':      ('228B22', 'FFFFFF'),
    'NCAAF':    ('8B0000', 'FFFFFF'),
    'NCAAB':    ('FF6347', 'FFFFFF'),
    'EPL':      ('4169E1', 'FFFFFF'),
    'MLS':      ('2E8B57', 'FFFFFF'),
    'OTHER':    ('808080', 'FFFFFF'),
}

BET_TYPE_COLORS = {
    'H2H':      ('CC3300', 'FFFFFF'),   # red-ish  (Moneyline)
    'SPREADS':  ('FF8800', 'FFFFFF'),   # orange   (Spread)
    'TOTALS':   ('0066CC', 'FFFFFF'),   # blue     (Over/Under)
    'HEDGE':    ('009999', 'FFFFFF'),   # teal
}

RESULT_COLORS = {
    'WIN':      ('00AA44', 'FFFFFF'),
    'LOSS':     ('CC0000', 'FFFFFF'),
    'VOID':     ('888888', 'FFFFFF'),
    'PENDING':  ('AAAAAA', '000000'),
}

BET_TYPE_LABELS = {
    'H2H':     'MONEYLINE',
    'SPREADS': 'SPREAD',
    'TOTALS':  'OVER/UNDER',
}


def _sport_short(sport_key):
    mapping = {
        'americanfootball_nfl':   'NFL',
        'americanfootball_ncaaf': 'NCAAF',
        'basketball_nba':         'NBA',
        'basketball_ncaab':       'NCAAB',
        'icehockey_nhl':          'NHL',
        'baseball_mlb':           'MLB',
        'soccer_epl':             'EPL',
        'soccer_mls':             'MLS',
    }
    return mapping.get(sport_key, 'OTHER')


def _badge_cell(ws, row, col, text, bg, fg):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=True, color=fg, size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = _border()
    return cell


def _plain_cell(ws, row, col, value, bold=False, align='center', num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = _border()
    if num_format:
        cell.number_format = num_format
    return cell


def generate_tracker():
    picks    = _get_all_picks_ordered()
    wb       = Workbook()
    ws       = wb.active
    ws.title = 'Picks Tracker'

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 20

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [12, 28, 14, 10, 12, 12, 10, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Headers ────────────────────────────────────────────────────────────────
    headers = ['DATE', 'GAME', 'SPORT TYPE', 'ODDS', 'UNITS BET',
               'BET AMOUNT', 'RESULT', 'CHANGE IN $', f'BANKROLL (start ${config.STARTING_BANKROLL:,.0f})']
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = _border()

    # ── Data rows ──────────────────────────────────────────────────────────────
    unit_value   = config.STARTING_BANKROLL * 0.01
    bankroll     = config.STARTING_BANKROLL

    for row_idx, pick in enumerate(picks, 2):
        ws.row_dimensions[row_idx].height = 18

        sport_short = _sport_short(pick['sport'])
        bet_label   = BET_TYPE_LABELS.get(pick['bet_type'], pick['bet_type'])
        result      = pick['result']
        units       = pick['units']
        odds        = pick['odds']
        bet_amount  = units * unit_value

        if result == 'WIN':
            change = units * unit_value * (odds - 1)
            bankroll += change
        elif result == 'LOSS':
            change = -(units * unit_value)
            bankroll += change
        else:
            change = 0.0

        snap_bankroll = bankroll if result != 'PENDING' else None

        # Date
        try:
            date_obj = datetime.strptime(pick['date'], '%Y-%m-%d').date()
        except Exception:
            date_obj = pick['date']
        _plain_cell(ws, row_idx, 1, date_obj, num_format='MM-DD-YY')

        # Game
        game_str = f"{pick['away_team']} @ {pick['home_team']}  |  {pick['bet_on']} ({pick['bet_label']})"
        _plain_cell(ws, row_idx, 2, game_str, align='left')

        # Sport badge
        sc = SPORT_COLORS.get(sport_short, ('808080', 'FFFFFF'))
        _badge_cell(ws, row_idx, 3, sport_short, sc[0], sc[1])

        # Odds
        _plain_cell(ws, row_idx, 4, round(odds, 2))

        # Units
        _plain_cell(ws, row_idx, 5, units)

        # Bet amount
        _plain_cell(ws, row_idx, 6, bet_amount, num_format='"$"#,##0.00')

        # Result badge
        rc = RESULT_COLORS.get(result, ('AAAAAA', '000000'))
        _badge_cell(ws, row_idx, 7, result, rc[0], rc[1])

        # Change in $
        change_cell = _plain_cell(ws, row_idx, 8,
                                  change if result != 'PENDING' else None,
                                  num_format='"$"#,##0.00;[Red]-"$"#,##0.00')
        if result == 'WIN':
            change_cell.font = _font(bold=True, color='00AA44')
        elif result == 'LOSS':
            change_cell.font = _font(bold=True, color='CC0000')

        # Bankroll snapshot
        _plain_cell(ws, row_idx, 9,
                    snap_bankroll,
                    bold=True,
                    num_format='"$"#,##0.00')

    # ── Summary row ────────────────────────────────────────────────────────────
    graded = [p for p in picks if p['result'] in ('WIN', 'LOSS')]
    wins   = sum(1 for p in graded if p['result'] == 'WIN')
    losses = sum(1 for p in graded if p['result'] == 'LOSS')
    wr     = (wins / len(graded) * 100) if graded else 0
    profit = bankroll - config.STARTING_BANKROLL

    summary_row = len(picks) + 3
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    summary = ws.cell(row=summary_row, column=1,
                      value=f'Record: {wins}W — {losses}L   |   Win Rate: {wr:.1f}%')
    summary.fill      = _fill('222222')
    summary.font      = _font(bold=True, color='FFFFFF', size=11)
    summary.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'D{summary_row}:F{summary_row}')
    profit_cell = ws.cell(row=summary_row, column=4,
                          value=f'P/L: {"+" if profit >= 0 else ""}${profit:,.2f}')
    profit_color = '00AA44' if profit >= 0 else 'CC0000'
    profit_cell.fill      = _fill('222222')
    profit_cell.font      = _font(bold=True, color=profit_color, size=11)
    profit_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f'G{summary_row}:I{summary_row}')
    bank_cell = ws.cell(row=summary_row, column=7,
                        value=f'Bankroll: ${bankroll:,.2f}')
    bank_cell.fill      = _fill('222222')
    bank_cell.font      = _font(bold=True, color='FFD700', size=11)
    bank_cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(TRACKER_PATH)
    return TRACKER_PATH


def _get_all_picks_ordered():
    import sqlite3
    from database import DB_PATH
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT * FROM picks ORDER BY date ASC, id ASC')
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows
